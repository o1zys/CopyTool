import os
import openpyxl.styles as sty
import openpyxl
import xlrd
import error
import global_var as gl
import shutil


def copy_xls(old_xls):
    new_xls = openpyxl.Workbook()
    old_sheet = old_xls.sheet_by_index(0)
    new_sheet = new_xls.active
    max_row = old_sheet.nrows
    max_col = old_sheet.ncols

    for m in range(0, max_row):
        # for n in range(97, 97+max_col):
        for n in range(0, max_col):
            integer = n // 26
            remainder = n % 26
            if integer == 0:
                n_str = chr(remainder+97)
            else:
                n_str = chr(integer+96) + chr(remainder+97)

            i = '%s%d' % (n_str, m+1)
            cell = old_sheet.cell_value(m, n)
            new_sheet[i].value = cell
    return new_xls


def execute(file_a, file_b, v1, v2):
    output_dir = file_a.replace('.xlsx', '')
    # read output file, xls
    try:
        xls_a = xlrd.open_workbook(file_a)
    except Exception as e:
        error.Error.set_code(3, str(e))
        return
    try:
        xls_b = xlrd.open_workbook(file_b)
    except Exception as e:
        error.Error.set_code(3, str(e))
        return
    # save old file
    try:
        shutil.copy(file_a, file_a.replace('.xlsx', '_old.xlsx'))
    except Exception as e:
        error.Error.set_code(6, str(e))
        return
    #old_sheet = old_xls.active
    old_sheet = xls_a.sheet_by_index(0)
    ref_sheet = xls_b.sheet_by_index(0)
    new_xls = copy_xls(xls_a)
    new_sheet = new_xls.active

    
    # construct dict shared id, dict value
    dict_sid = {}
    dict_val = {}
    dict_row = {}
    dict_lang_key = {}
    map_lang_col = {}
    col_id = col_sid = col_inst = col_ignore = col_feature = col_langkey = col_desc = col_hist = col_term = None
    skip_row = True
    for i in range(old_sheet.nrows):
        # read number id
        if i == 0:
            for j in range(old_sheet.ncols):
                # feature, term, ignore, desc, inst, hist, langkey
                value = old_sheet.cell_value(i, j).lower()
                if value == "sheetid":
                    col_id = j
                elif value == "shareid":
                    col_sid = j
                elif value == "ignore":
                    col_ignore = j
                elif value == "history":
                    col_hist = j
                elif value == "feature":
                    col_feature = j
                elif value == "term":
                    col_term = j
                elif value == "description":
                    col_desc = j
                elif value == "instruction":
                    col_inst = j
                elif value == "langkey":
                    col_langkey = j
            for j in range(old_sheet.ncols):
                if j > col_langkey:
                    map_lang_col[old_sheet.cell_value(i, j)] = j
            continue
        if i == 1:
            sp_char = '#'
            if sp_char in old_sheet.cell_value(i, 0):
                skip_row = False
        if skip_row is True and i < gl.trans_content_row:
            continue
        tmp_id = old_sheet.cell_value(i, col_id)
        tmp_lang_key = old_sheet.cell_value(i, col_langkey)
        dict_sid[tmp_id] = old_sheet.cell_value(i, col_sid)
        dict_val[tmp_id] = tmp_lang_key
        dict_row[tmp_id] = i
        if tmp_lang_key != "":
            dict_lang_key[tmp_lang_key] = tmp_id

    # read ref_sheet
    tar_map_lang_col = {}
    skip_ref_row = True
    ref_col_id = ref_col_sid = ref_col_inst = ref_col_ignore = ref_col_feature = ref_col_langkey = ref_col_desc \
        = ref_col_hist = ref_col_term = None
    for i in range(ref_sheet.nrows):
        if i == 0:
            for j in range(ref_sheet.ncols):
                # feature, term, ignore, desc, inst, hist, langkey
                value = ref_sheet.cell_value(i, j).lower()
                if value == "sheetid":
                    ref_col_id = j
                elif value == "shareid":
                    ref_col_sid = j
                elif value == "ignore":
                    ref_col_ignore = j
                elif value == "history":
                    ref_col_hist = j
                elif value == "feature":
                    ref_col_feature = j
                elif value == "term":
                    ref_col_term = j
                elif value == "description":
                    ref_col_desc = j
                elif value == "instruction":
                    ref_col_inst = j
                elif value == "langkey":
                    ref_col_langkey = j
            for j in range(ref_sheet.ncols):
                if j > ref_col_langkey:
                    tar_map_lang_col[ref_sheet.cell_value(i, j)] = j
            continue
        if i == 1:
            sp_char = '#'
            if sp_char in ref_sheet.cell_value(i, 0):
                skip_ref_row = False
        if skip_ref_row is True and i < gl.trans_content_row:
            continue
        tar_id = ref_sheet.cell_value(i, ref_col_id)
        tar_val = ref_sheet.cell_value(i, ref_col_langkey)
        fill_type = sty.PatternFill(fill_type='solid', fgColor=gl.color_copy_modify)
        # special rule for hist
        if v2 == 0 and ref_col_hist is not None and col_hist is not None and ref_col_sid is not None and dict_val.__contains__(tar_id):
            if dict_val[tar_id] == tar_val and dict_sid[tar_id] == ref_sheet.cell_value(i, ref_col_sid):
                orig_hist = old_sheet.cell_value(dict_row[tar_id], col_hist)
                tar_hist = ref_sheet.cell_value(i, ref_col_hist)
                if tar_hist != orig_hist:
                    new_sheet.cell(row=dict_row[tar_id]+1, column=col_hist+1, value=tar_hist).fill \
                        = fill_type

        if ref_sheet.cell_value(i, col_sid) != "":
            continue
        is_fill_content = False
        if v1 == 0 and dict_lang_key.__contains__(tar_val):
            fill_type = sty.PatternFill(fill_type='solid', fgColor=gl.color_copy_modify)
            tar_id = dict_lang_key[tar_val]
            for key, val in tar_map_lang_col.items():
                if map_lang_col.__contains__(key):
                    orig_content = old_sheet.cell_value(dict_row[tar_id], map_lang_col[key])
                    tar_content = ref_sheet.cell_value(i, val)
                    if tar_content != "" and tar_content != orig_content:
                        new_sheet.cell(dict_row[tar_id]+1, map_lang_col[key]+1, tar_content).fill = fill_type
            is_fill_content = True
        elif v1 == 1 and dict_val.__contains__(tar_id):
                if dict_val[tar_id] != tar_val and dict_sid[tar_id] == "":
                    fill_type = sty.PatternFill(fill_type='solid', fgColor=gl.color_copy_unique)
                    for key, val in tar_map_lang_col.items():
                        if key == "CNS":
                            continue
                        if map_lang_col.__contains__(key):
                            orig_content = old_sheet.cell_value(dict_row[tar_id], map_lang_col[key])
                            tar_content = ref_sheet.cell_value(i, val)
                            if tar_content != "" and tar_content != orig_content:
                                new_sheet.cell(dict_row[tar_id]+1, map_lang_col[key]+1, tar_content).fill = fill_type
                    is_fill_content = True

        if v2 == 0 and is_fill_content:
            if ref_col_feature is not None and col_feature is not None:
                orig_feature = old_sheet.cell_value(dict_row[tar_id], col_feature)
                tar_feature = ref_sheet.cell_value(i, ref_col_feature)
                if tar_feature != orig_feature:
                    new_sheet.cell(row=dict_row[tar_id]+1, column=col_feature+1, value=tar_feature).fill \
                        = fill_type
            if ref_col_term is not None and col_term is not None:
                orig_term = old_sheet.cell_value(dict_row[tar_id], col_term)
                tar_term = ref_sheet.cell_value(i, ref_col_term)
                if tar_term != orig_term:
                    new_sheet.cell(row=dict_row[tar_id]+1, column=col_term+1, value=tar_term).fill \
                        = fill_type
            if ref_col_ignore is not None and col_ignore is not None:
                orig_ignore = old_sheet.cell_value(dict_row[tar_id], col_ignore)
                tar_ignore = ref_sheet.cell_value(i, ref_col_ignore)
                if tar_ignore != orig_ignore:
                    new_sheet.cell(row=dict_row[tar_id]+1, column=col_ignore+1, value=tar_ignore).fill \
                        = fill_type
            if ref_col_desc is not None and col_desc is not None:
                orig_desc = old_sheet.cell_value(dict_row[tar_id], col_desc)
                tar_desc = ref_sheet.cell_value(i, ref_col_desc)
                if tar_desc != orig_desc:
                    new_sheet.cell(row=dict_row[tar_id]+1, column=col_desc+1, value=tar_desc).fill\
                        = fill_type
            if ref_col_inst is not None and col_inst is not None:
                orig_inst = old_sheet.cell_value(dict_row[tar_id], col_inst)
                tar_inst = ref_sheet.cell_value(i, ref_col_inst)
                if tar_inst != orig_inst:
                    new_sheet.cell(row=dict_row[tar_id]+1, column=col_inst+1, value=tar_inst).fill\
                        = fill_type

    new_sheet.freeze_panes = "A2"
    try:
        os.remove(file_a)
    except Exception as e:
        error.Error.set_code(2, str(e))
        return

    try:
        new_xls.save(file_a)
    except Exception as e:
        error.Error.set_code(7, str(e))
        return

    new_xls.close()

